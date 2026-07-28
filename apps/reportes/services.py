import csv
import datetime
import re
from io import BytesIO, StringIO
from pathlib import Path

from django.conf import settings
from django.db.models import Count, Q
from django.utils import timezone
from django.utils.dateparse import parse_date

from apps.core.models import CentroSalud
from apps.dashboard.models import AlertaClinica
from apps.pacientes.models import Paciente
from apps.referencias.models import ReferenciaMedica
from apps.seguimiento.models import SeguimientoPaciente

from .models import ReporteGenerado


DOCUMENT_VERSION = '1.0'
DOCUMENT_DATE = '2026-06-23'
def _logo_path() -> str:
    from apps.core.utils import get_system_logo_path_or_fallback
    return get_system_logo_path_or_fallback()

REPORT_TYPES = [
    {
        'id': 'resumen_mensual',
        'titulo': 'Resumen Mensual',
        'descripcion': 'Estadisticas generales del periodo seleccionado',
        'icono': 'calendar_month',
        'modelo': ReporteGenerado.TipoReporte.RESUMEN_MENSUAL,
    },
    {
        'id': 'por_provincia',
        'titulo': 'Por Provincia',
        'descripcion': 'Distribucion geografica de pacientes, centros y referencias',
        'icono': 'map',
        'modelo': ReporteGenerado.TipoReporte.POR_PROVINCIA,
    },
    {
        'id': 'por_diagnostico',
        'titulo': 'Por Diagnostico',
        'descripcion': 'Frecuencia y porcentaje por diagnostico',
        'icono': 'medical_information',
        'modelo': ReporteGenerado.TipoReporte.POR_DIAGNOSTICO,
    },
    {
        'id': 'seguimiento',
        'titulo': 'Seguimiento de Casos',
        'descripcion': 'Estado actual, fase, responsable y alertas pendientes',
        'icono': 'monitor_heart',
        'modelo': ReporteGenerado.TipoReporte.SEGUIMIENTO_CASOS,
    },
    {
        'id': 'referencias',
        'titulo': 'Referencias Medicas',
        'descripcion': 'Referencias enviadas, recibidas, pendientes y criticas',
        'icono': 'send',
        'modelo': ReporteGenerado.TipoReporte.REFERENCIAS_MEDICAS,
    },
]

REPORT_TYPE_MAP = {item['id']: item for item in REPORT_TYPES}

FORMAT_CHOICES = [
    {'id': ReporteGenerado.Formato.PDF, 'label': 'PDF', 'icono': 'picture_as_pdf'},
    {'id': ReporteGenerado.Formato.XLSX, 'label': 'Excel .xlsx', 'icono': 'table_chart'},
    {'id': ReporteGenerado.Formato.CSV, 'label': 'CSV', 'icono': 'csv'},
]

PERIOD_CHOICES = [
    {'id': 'ultimo_mes', 'label': 'Ultimo mes'},
    {'id': 'ultimos_3_meses', 'label': 'Ultimos 3 meses'},
    {'id': 'ultimos_6_meses', 'label': 'Ultimos 6 meses'},
    {'id': 'ultimo_anio', 'label': 'Ultimo ano'},
    {'id': 'personalizado', 'label': 'Rango personalizado'},
]

PERIOD_LABELS = {item['id']: item['label'] for item in PERIOD_CHOICES}


def _hace_n_meses(fecha, n):
    mes = fecha.month - n
    anio = fecha.year
    while mes <= 0:
        mes += 12
        anio -= 1
    ultimo_dia = [
        31,
        29 if anio % 4 == 0 and (anio % 100 != 0 or anio % 400 == 0) else 28,
        31, 30, 31, 30, 31, 31, 30, 31, 30, 31,
    ][mes - 1]
    return fecha.replace(year=anio, month=mes, day=min(fecha.day, ultimo_dia))


def _nombre_usuario(user):
    if not user:
        return 'Usuario no disponible'
    nombre = getattr(user, 'nombre_completo', '') or user.get_full_name()
    return nombre or user.get_username()


def _choice_label(choices, value, default='Sin especificar'):
    if not value:
        return default
    return dict(choices).get(value, value)


def _fmt_date(value):
    if not value:
        return '-'
    if hasattr(value, 'date') and not isinstance(value, datetime.date):
        value = timezone.localtime(value).date()
    return value.strftime('%d/%m/%Y')


def _fmt_datetime(value):
    if not value:
        return '-'
    if timezone.is_aware(value):
        value = timezone.localtime(value)
    return value.strftime('%d/%m/%Y %H:%M')


def _pct(value, total):
    if not total:
        return '0%'
    return f'{round((value / total) * 100, 1)}%'


def parse_report_params(data, user=None):
    today = timezone.localdate()
    tipo_id = (data.get('tipo_reporte') or '').strip()
    formato = (data.get('formato') or '').strip().lower()
    periodo = (data.get('periodo') or 'ultimo_mes').strip()
    provincia = (data.get('provincia') or '').strip()
    medico = (data.get('medico') or '').strip()

    errors = []
    if tipo_id not in REPORT_TYPE_MAP:
        errors.append('Seleccione un tipo de reporte valido.')
    if formato not in ReporteGenerado.Formato.values:
        errors.append('Seleccione un formato valido.')
    if periodo not in PERIOD_LABELS:
        errors.append('Seleccione un periodo valido.')

    if periodo == 'personalizado':
        fecha_inicio = parse_date(data.get('fecha_inicio', ''))
        fecha_fin = parse_date(data.get('fecha_fin', ''))
        if not fecha_inicio or not fecha_fin:
            errors.append('Ingrese fecha desde y fecha hasta para el rango personalizado.')
    else:
        meses = {
            'ultimo_mes': 1,
            'ultimos_3_meses': 3,
            'ultimos_6_meses': 6,
            'ultimo_anio': 12,
        }[periodo]
        fecha_inicio = _hace_n_meses(today, meses)
        fecha_fin = today

    if fecha_inicio and fecha_fin and fecha_inicio > fecha_fin:
        errors.append('La fecha desde no puede ser mayor que la fecha hasta.')

    return {
        'errors': errors,
        'tipo_id': tipo_id,
        'formato': formato,
        'periodo': periodo,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'provincia': provincia,
        'medico': medico,
        'user': user,
    }


def _pacientes_filtrados(params, aplicar_periodo=True):
    qs = Paciente.objects.select_related('medico_asignado')
    if aplicar_periodo:
        qs = qs.filter(
            created_at__date__gte=params['fecha_inicio'],
            created_at__date__lte=params['fecha_fin'],
        )
    if params.get('provincia'):
        qs = qs.filter(provincia=params['provincia'])
    if params.get('medico'):
        qs = qs.filter(medico_asignado_id=params['medico'])
    return qs


def _referencias_filtradas(params):
    qs = ReferenciaMedica.objects.select_related(
        'paciente', 'medico_referente', 'especialista_destino'
    ).filter(
        fecha_referencia__date__gte=params['fecha_inicio'],
        fecha_referencia__date__lte=params['fecha_fin'],
    )
    if params.get('provincia'):
        qs = qs.filter(paciente__provincia=params['provincia'])
    if params.get('medico'):
        qs = qs.filter(
            Q(medico_referente_id=params['medico']) |
            Q(especialista_destino_id=params['medico'])
        )
    return qs


def _seguimientos_filtrados(params):
    qs = SeguimientoPaciente.objects.select_related('paciente', 'medico').filter(
        fecha_seguimiento__date__gte=params['fecha_inicio'],
        fecha_seguimiento__date__lte=params['fecha_fin'],
    )
    if params.get('provincia'):
        qs = qs.filter(paciente__provincia=params['provincia'])
    if params.get('medico'):
        qs = qs.filter(medico_id=params['medico'])
    return qs


def _base_report(params, titulo, columns, rows, summary=None, observaciones=''):
    now = timezone.localtime()
    tipo = REPORT_TYPE_MAP[params['tipo_id']]
    periodo_label = (
        f"{PERIOD_LABELS[params['periodo']]} "
        f"({_fmt_date(params['fecha_inicio'])} - {_fmt_date(params['fecha_fin'])})"
    )
    codigo = f"FACCI-REP-{now:%Y%m%d}-{tipo['modelo']}"
    return {
        'tipo_id': params['tipo_id'],
        'tipo_modelo': tipo['modelo'],
        'titulo': titulo,
        'codigo_documento': codigo,
        'version': DOCUMENT_VERSION,
        'fecha_documento': DOCUMENT_DATE,
        'fecha_generacion': now,
        'fecha_inicio': params['fecha_inicio'],
        'fecha_fin': params['fecha_fin'],
        'periodo_label': periodo_label,
        'generado_por': _nombre_usuario(params.get('user')),
        'columns': columns,
        'rows': rows,
        'summary': summary or [],
        'observaciones': observaciones,
        'logo_path': _logo_path(),
        'logo_static': 'img/Logo-Facci-1.png',
        'total_registros': len(rows),
    }


def _build_resumen_mensual(params):
    pacientes_periodo = _pacientes_filtrados(params)
    pacientes_total = _pacientes_filtrados(params, aplicar_periodo=False)
    activos = pacientes_total.filter(
        estado_actual__in=[
            Paciente.EstadoPaciente.SOSPECHOSO,
            Paciente.EstadoPaciente.REFERIDO,
            Paciente.EstadoPaciente.EN_ESTUDIO,
            Paciente.EstadoPaciente.CONFIRMADO,
            Paciente.EstadoPaciente.EN_TRATAMIENTO,
        ]
    ).count()
    alertas = AlertaClinica.objects.filter(
        fecha_creacion__date__gte=params['fecha_inicio'],
        fecha_creacion__date__lte=params['fecha_fin'],
    )
    referencias = _referencias_filtradas(params)
    seguimientos = _seguimientos_filtrados(params)

    rows = [
        {'indicador': 'Total pacientes registrados', 'valor': pacientes_total.count()},
        {'indicador': 'Pacientes activos', 'valor': activos},
        {'indicador': 'Nuevos pacientes en periodo', 'valor': pacientes_periodo.count()},
        {'indicador': 'Alertas generadas', 'valor': alertas.count()},
        {'indicador': 'Referencias creadas', 'valor': referencias.count()},
        {'indicador': 'Seguimientos registrados', 'valor': seguimientos.count()},
    ]
    return _base_report(
        params,
        'Resumen Mensual',
        [('indicador', 'Indicador'), ('valor', 'Valor')],
        rows,
        summary=[
            {'label': 'Pacientes', 'value': pacientes_total.count()},
            {'label': 'Nuevos', 'value': pacientes_periodo.count()},
            {'label': 'Alertas', 'value': alertas.count()},
            {'label': 'Referencias', 'value': referencias.count()},
        ],
    )


def _build_por_provincia(params):
    pacientes = _pacientes_filtrados(params, aplicar_periodo=False)
    if params.get('provincia'):
        provincias = [params['provincia']]
    else:
        provincias = list(
            set(pacientes.values_list('provincia', flat=True)) |
            set(CentroSalud.objects.values_list('provincia', flat=True))
        )
    rows = []
    referencias = _referencias_filtradas(params)
    for provincia in sorted(p for p in provincias if p):
        rows.append({
            'provincia': provincia,
            'pacientes': pacientes.filter(provincia=provincia).count(),
            'centros': CentroSalud.objects.filter(provincia=provincia).count(),
            'referencias': referencias.filter(paciente__provincia=provincia).count(),
        })
    rows.sort(key=lambda item: item['pacientes'], reverse=True)
    return _base_report(
        params,
        'Reporte por Provincia',
        [
            ('provincia', 'Provincia'),
            ('pacientes', 'Pacientes'),
            ('centros', 'Centros de salud'),
            ('referencias', 'Referencias del periodo'),
        ],
        rows,
        summary=[
            {'label': 'Provincias', 'value': len(rows)},
            {'label': 'Pacientes', 'value': sum(row['pacientes'] for row in rows)},
            {'label': 'Centros', 'value': sum(row['centros'] for row in rows)},
        ],
    )


def _build_por_diagnostico(params):
    pacientes = _pacientes_filtrados(params)
    total = pacientes.count()
    agrupados = pacientes.values('diagnostico').annotate(cantidad=Count('id')).order_by('-cantidad')
    rows = []
    for item in agrupados:
        cantidad = item['cantidad']
        rows.append({
            'diagnostico': _choice_label(Paciente.Diagnostico.choices, item['diagnostico']),
            'cantidad': cantidad,
            'porcentaje': _pct(cantidad, total),
        })
    return _base_report(
        params,
        'Reporte por Diagnostico',
        [
            ('diagnostico', 'Diagnostico'),
            ('cantidad', 'Cantidad'),
            ('porcentaje', 'Porcentaje'),
        ],
        rows,
        summary=[
            {'label': 'Diagnosticos', 'value': len(rows)},
            {'label': 'Pacientes en periodo', 'value': total},
        ],
    )


def _build_seguimiento(params):
    seguimientos = list(_seguimientos_filtrados(params)[:500])
    rows = []
    for seguimiento in seguimientos:
        paciente = seguimiento.paciente
        alertas_pendientes = paciente.alertas_clinicas.filter(
            estado__in=[AlertaClinica.Estado.PENDIENTE, AlertaClinica.Estado.REVISADA]
        ).count()
        rows.append({
            'paciente': paciente.nombre_completo,
            'estado_actual': paciente.get_estado_actual_display(),
            'fase': seguimiento.get_fase_protocolo_display(),
            'ultima_actualizacion': _fmt_datetime(seguimiento.updated_at),
            'responsable': _nombre_usuario(seguimiento.medico),
            'alertas_pendientes': alertas_pendientes,
        })
    return _base_report(
        params,
        'Seguimiento de Casos',
        [
            ('paciente', 'Paciente'),
            ('estado_actual', 'Estado actual'),
            ('fase', 'Fase'),
            ('ultima_actualizacion', 'Ultima actualizacion'),
            ('responsable', 'Responsable'),
            ('alertas_pendientes', 'Alertas pendientes'),
        ],
        rows,
        summary=[
            {'label': 'Seguimientos', 'value': len(rows)},
            {'label': 'Con alertas', 'value': sum(1 for row in rows if row['alertas_pendientes'])},
        ],
    )


def _build_referencias(params):
    referencias = list(_referencias_filtradas(params)[:700])
    rows = []
    for referencia in referencias:
        rows.append({
            'codigo': referencia.codigo,
            'paciente': referencia.paciente.nombre_completo,
            'estado': referencia.get_estado_display(),
            'prioridad': referencia.get_prioridad_display(),
            'centro_origen': _nombre_usuario(referencia.medico_referente),
            'centro_destino': referencia.hospital_destino.nombre if referencia.hospital_destino else '',
            'fecha': _fmt_datetime(referencia.fecha_referencia),
        })
    pendientes = sum(1 for ref in referencias if ref.estado == ReferenciaMedica.EstadoReferencia.PENDIENTE)
    atendidas = sum(1 for ref in referencias if ref.estado == ReferenciaMedica.EstadoReferencia.COMPLETADA)
    criticas = sum(1 for ref in referencias if ref.prioridad in [ReferenciaMedica.Prioridad.ALTA, ReferenciaMedica.Prioridad.URGENTE])
    return _base_report(
        params,
        'Referencias Medicas',
        [
            ('codigo', 'Codigo'),
            ('paciente', 'Paciente'),
            ('estado', 'Estado'),
            ('prioridad', 'Prioridad'),
            ('centro_origen', 'Origen'),
            ('centro_destino', 'Destino'),
            ('fecha', 'Fecha'),
        ],
        rows,
        summary=[
            {'label': 'Enviadas', 'value': len(rows)},
            {'label': 'Recibidas', 'value': sum(1 for ref in referencias if ref.especialista_destino_id)},
            {'label': 'Pendientes', 'value': pendientes},
            {'label': 'Atendidas', 'value': atendidas},
            {'label': 'Criticas', 'value': criticas},
        ],
    )


REPORT_BUILDERS = {
    'resumen_mensual': _build_resumen_mensual,
    'por_provincia': _build_por_provincia,
    'por_diagnostico': _build_por_diagnostico,
    'seguimiento': _build_seguimiento,
    'referencias': _build_referencias,
}


def build_report(params):
    return REPORT_BUILDERS[params['tipo_id']](params)


def _safe_filename(value):
    value = re.sub(r'[^a-zA-Z0-9_-]+', '_', value.lower()).strip('_')
    return value or 'reporte'


def report_filename(report, formato):
    ext = 'xlsx' if formato == ReporteGenerado.Formato.XLSX else formato
    return f"{_safe_filename(report['titulo'])}_{timezone.localtime():%Y%m%d_%H%M}.{ext}"


def _plain(value):
    return str(value if value is not None else '')


def render_csv(report):
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['FACCI Care', report['titulo']])
    writer.writerow(['Codigo', report['codigo_documento']])
    writer.writerow(['Version', report['version']])
    writer.writerow(['Fecha', _fmt_datetime(report['fecha_generacion'])])
    writer.writerow(['Periodo', report['periodo_label']])
    writer.writerow(['Generado por', report['generado_por']])
    writer.writerow([])
    if report['summary']:
        writer.writerow(['Resumen'])
        for item in report['summary']:
            writer.writerow([item['label'], item['value']])
        writer.writerow([])
    writer.writerow([label for _, label in report['columns']])
    if report['rows']:
        for row in report['rows']:
            writer.writerow([_plain(row.get(key, '')) for key, _ in report['columns']])
    else:
        writer.writerow(['No hay registros para el periodo seleccionado'])
    return output.getvalue().encode('utf-8-sig')


def render_excel(report):
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = report['titulo'][:31]
    ws.sheet_view.showGridLines = False

    row = 1
    if report['logo_path']:
        try:
            logo = ExcelImage(report['logo_path'])
            logo.width = 120
            logo.height = 65
            ws.add_image(logo, 'A1')
        except Exception:
            pass
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row=row, column=2, value='FACCI Care').font = Font(size=16, bold=True, color='2E6F73')
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row=row, column=2, value=report['titulo']).font = Font(size=14, bold=True)
    row += 2

    meta = [
        ('Codigo del documento', report['codigo_documento']),
        ('Version', report['version']),
        ('Fecha documento', report['fecha_documento']),
        ('Periodo', report['periodo_label']),
        ('Fecha de generacion', _fmt_datetime(report['fecha_generacion'])),
        ('Generado por', report['generado_por']),
    ]
    for label, value in meta:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1
    row += 1

    if report['summary']:
        ws.cell(row=row, column=1, value='Resumen').font = Font(bold=True, color='FFFFFF')
        ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor='2E6F73')
        row += 1
        for item in report['summary']:
            ws.cell(row=row, column=1, value=item['label']).font = Font(bold=True)
            ws.cell(row=row, column=2, value=item['value'])
            row += 1
        row += 1

    headers = [label for _, label in report['columns']]
    border = Border(
        left=Side(style='thin', color='B8C7C9'),
        right=Side(style='thin', color='B8C7C9'),
        top=Side(style='thin', color='B8C7C9'),
        bottom=Side(style='thin', color='B8C7C9'),
    )
    header_fill = PatternFill('solid', fgColor='D7ECEE')
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color='1F4F54')
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    row += 1

    if report['rows']:
        for item in report['rows']:
            for col, (key, _) in enumerate(report['columns'], start=1):
                cell = ws.cell(row=row, column=col, value=_plain(item.get(key, '')))
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            row += 1
    else:
        ws.cell(row=row, column=1, value='No hay registros para el periodo seleccionado')

    for col in range(1, len(headers) + 1):
        max_len = 12
        for cell in ws[get_column_letter(col)]:
            max_len = max(max_len, len(str(cell.value or '')) + 2)
        ws.column_dimensions[get_column_letter(col)].width = min(max_len, 36)

    resumen = wb.create_sheet('Resumen')
    resumen.append(['Campo', 'Valor'])
    resumen.append(['Reporte', report['titulo']])
    resumen.append(['Codigo', report['codigo_documento']])
    resumen.append(['Periodo', report['periodo_label']])
    resumen.append(['Generado por', report['generado_por']])
    resumen.append(['Total registros', report['total_registros']])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def render_pdf(report):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=0.3 * inch,
        leftMargin=0.3 * inch,
        topMargin=0.28 * inch,
        bottomMargin=0.38 * inch,
        title=report['titulo'],
    )
    styles = getSampleStyleSheet()
    styles['Title'].fontSize = 13.5
    styles['Title'].leading = 15.5
    styles['Title'].textColor = colors.HexColor('#1F4F54')
    styles['Normal'].fontSize = 6.8
    styles['Normal'].leading = 8.2

    story = []
    header_cells = []
    if report['logo_path']:
        try:
            header_cells.append(Image(report['logo_path'], width=0.9 * inch, height=0.58 * inch))
        except Exception:
            header_cells.append(Paragraph('FACCI', styles['Title']))
    else:
        header_cells.append(Paragraph('FACCI', styles['Title']))
    header_cells.append(Paragraph(f"<b>{report['titulo']}</b><br/>Formato institucional FACCI", styles['Title']))
    header_cells.append(Paragraph(
        f"<b>Codigo:</b> {report['codigo_documento']}<br/>"
        f"<b>Version:</b> {report['version']}<br/>"
        f"<b>Fecha:</b> {report['fecha_documento']}",
        styles['Normal'],
    ))
    header_table = Table([header_cells], colWidths=[1.05 * inch, 4.55 * inch, 1.9 * inch])
    header_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.8, colors.HexColor('#2E6F73')),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#B8C7C9')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F6FBFB')),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.06 * inch))

    meta = [
        ['Reporte', report['titulo'], 'Periodo', report['periodo_label']],
        ['Generado por', report['generado_por'], 'Generado el', _fmt_datetime(report['fecha_generacion'])],
    ]
    meta_table = Table(meta, colWidths=[1.1 * inch, 2.5 * inch, 1.1 * inch, 2.6 * inch])
    meta_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#B8C7C9')),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#D7ECEE')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#D7ECEE')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 6.7),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 0.06 * inch))

    if report['summary']:
        summary_data = [['Indicador', 'Valor']] + [[item['label'], item['value']] for item in report['summary']]
        summary_table = Table(summary_data, colWidths=[3.3 * inch, 1.2 * inch])
        summary_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#B8C7C9')),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E6F73')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 6.7),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]))
        story.append(Paragraph('<b>Resumen</b>', styles['Normal']))
        story.append(summary_table)
        story.append(Spacer(1, 0.06 * inch))

    table_data = [[label for _, label in report['columns']]]
    if report['rows']:
        for row in report['rows']:
            table_data.append([
                Paragraph(_plain(row.get(key, '')), styles['Normal'])
                for key, _ in report['columns']
            ])
    else:
        table_data.append([Paragraph('No hay registros para el periodo seleccionado', styles['Normal'])] + [''] * (len(report['columns']) - 1))

    usable_width = 7.5 * inch
    col_width = usable_width / max(len(report['columns']), 1)
    data_table = Table(table_data, repeatRows=1, colWidths=[col_width] * len(report['columns']))
    data_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#B8C7C9')),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E6F73')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 6.2),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F6FBFB')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 1.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1.5),
    ]))
    story.append(Paragraph('<b>Resultados</b>', styles['Normal']))
    story.append(data_table)

    if report['observaciones']:
        story.append(Spacer(1, 0.06 * inch))
        story.append(Paragraph(f"<b>Observaciones:</b> {report['observaciones']}", styles['Normal']))

    def footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setFont('Helvetica', 6.5)
        canvas.setFillColor(colors.HexColor('#4A5F61'))
        canvas.drawString(0.3 * inch, 0.24 * inch, 'FACCI Care - Fundacion Amigos Contra el Cancer Infantil')
        canvas.drawCentredString(4.25 * inch, 0.24 * inch, f"Pagina {doc_obj.page}")
        canvas.drawRightString(8.2 * inch, 0.24 * inch, report['codigo_documento'])
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    return buffer.getvalue()


def render_report_file(report, formato):
    if formato == ReporteGenerado.Formato.PDF:
        return render_pdf(report), 'application/pdf'
    if formato == ReporteGenerado.Formato.XLSX:
        return render_excel(report), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return render_csv(report), 'text/csv; charset=utf-8'
